import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_pothys_template():
    wb = openpyxl.Workbook()
    
    # Common Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="D4AF37")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    label_font = Font(name=font_family, size=11, bold=True, color="000000")
    regular_font = Font(name=font_family, size=11)
    bold_font = Font(name=font_family, size=11, bold=True)
    
    fill_gold = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    fill_dark = PatternFill(start_color="1A1A22", end_color="1A1A22", fill_type="solid")
    fill_light_gold = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ────────────────────────────────────────────────────────
    # SHEET 1: Branch Summary
    # ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Branch Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A1:D2")
    title_cell = ws1["A1"]
    title_cell.value = "POTHYS SWARNA MAHAL - DAILY PERFORMANCE REPORT"
    title_cell.font = title_font
    title_cell.fill = fill_dark
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata Block
    metadata = [
        ("Report Date", "2026-07-17", "Branch Name", "Coimbatore Swarna Mahal"),
        ("Manager Name", "Manager Name", "Sub Manager Name", "Sub Manager Name")
    ]
    for row_idx, row_data in enumerate(metadata, start=4):
        ws1.cell(row=row_idx, column=1, value=row_data[0]).font = label_font
        ws1.cell(row=row_idx, column=2, value=row_data[1]).font = regular_font
        ws1.cell(row=row_idx, column=3, value=row_data[2]).font = label_font
        ws1.cell(row=row_idx, column=4, value=row_data[3]).font = regular_font
        
        for col in range(1, 5):
            cell = ws1.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.fill = fill_light_gold
            
    # Section: Sales Performance
    ws1.cell(row=7, column=1, value="SALES PERFORMANCE (INR)").font = section_font
    sales_metrics = [
        ("Gold Sales (Amount)", 0.0),
        ("Silver Sales (Amount)", 0.0),
        ("Platinum Sales (Amount)", 0.0),
        ("Diamond Sales (Amount)", 0.0),
        ("Total Revenue", "=SUM(B8:B11)")
    ]
    for idx, (label, val) in enumerate(sales_metrics, start=8):
        ws1.cell(row=idx, column=1, value=label).font = label_font if label == "Total Revenue" else regular_font
        val_cell = ws1.cell(row=idx, column=2, value=val)
        val_cell.font = bold_font if label == "Total Revenue" else regular_font
        if label != "Total Revenue":
            val_cell.number_format = '₹#,##,##0.00'
        else:
            val_cell.number_format = '₹#,##,##0.00'
            val_cell.fill = fill_light_gold
        
        ws1.cell(row=idx, column=1).border = thin_border
        val_cell.border = thin_border

    # Section: Scheme Enrollments
    ws1.cell(row=7, column=3, value="SCHEME ENROLLMENTS").font = section_font
    schemes = [
        ("DigiGold Enrollments", 0),
        ("DigiSilver Enrollments", 0)
    ]
    for idx, (label, val) in enumerate(schemes, start=8):
        ws1.cell(row=idx, column=3, value=label).font = regular_font
        val_cell = ws1.cell(row=idx, column=4, value=val)
        val_cell.font = regular_font
        
        ws1.cell(row=idx, column=3).border = thin_border
        val_cell.border = thin_border

    # Section: Human Resources
    ws1.cell(row=11, column=3, value="HUMAN RESOURCES").font = section_font
    hr = [
        ("Employees Present", 0),
        ("Employees Absent", 0)
    ]
    for idx, (label, val) in enumerate(hr, start=12):
        ws1.cell(row=idx, column=3, value=label).font = regular_font
        val_cell = ws1.cell(row=idx, column=4, value=val)
        val_cell.font = regular_font
        
        ws1.cell(row=idx, column=3).border = thin_border
        val_cell.border = thin_border

    # Section: Operations & Remarks
    ws1.cell(row=14, column=1, value="OPERATIONAL SUMMARY").font = section_font
    ops = [
        ("Customer Complaints", "None"),
        ("Operational Issues", "None"),
        ("Manager Remarks", "All operations running smoothly.")
    ]
    for idx, (label, val) in enumerate(ops, start=15):
        ws1.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
        ws1.cell(row=idx, column=1, value=label).font = label_font
        val_cell = ws1.cell(row=idx, column=2, value=val)
        val_cell.font = regular_font
        
        # Border application for merged cells
        for col in range(1, 5):
            ws1.cell(row=idx, column=col).border = thin_border

    # Column width formatting
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 25)

    # ────────────────────────────────────────────────────────
    # SHEET 2: Employee Performance
    # ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Employee Performance")
    ws2.views.sheetView[0].showGridLines = True
    
    headers = [
        "Employee Name", "Designation", "Gold Grams Sold", "Gold Amount",
        "Silver Grams Sold", "Silver Amount", "Platinum Amount", "Diamond Amount",
        "DigiGold Enrollments", "DigiSilver Enrollments"
    ]
    
    # Set headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_gold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Mock row data placeholder
    mock_employees = [
        ("Rajesh Kumar", "Sales Executive", 12.5, 95000.00, 150.0, 18000.00, 15000.00, 45000.00, 5, 8),
        ("Priya Sharma", "Sales Executive", 8.2, 62000.00, 80.0, 9600.00, 8000.00, 0.00, 3, 4),
        ("Anand Singh", "Senior Executive", 20.0, 152000.00, 300.0, 36000.00, 25000.00, 95000.00, 12, 15)
    ]
    for row_idx, emp in enumerate(mock_employees, start=2):
        for col_idx, val in enumerate(emp, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            # formatting amounts
            if col_idx in [4, 6, 7, 8]:
                cell.number_format = '₹#,##,##0.00'
            elif col_idx in [3, 5]:
                cell.number_format = '#,##0.00'

    # Column width formatting
    ws2.row_dimensions[1].height = 28
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ────────────────────────────────────────────────────────
    # SHEET 3: Scheme Summary
    # ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Scheme Summary")
    ws3.views.sheetView[0].showGridLines = True
    
    # Title
    ws3.merge_cells("A1:C1")
    title3 = ws3["A1"]
    title3.value = "DAILY SCHEME SUMMARY"
    title3.font = title_font
    title3.fill = fill_dark
    title3.alignment = Alignment(horizontal="center", vertical="center")
    
    scheme_summary = [
        ("DigiGold Total", "=SUM('Employee Performance'!I2:I100)"),
        ("DigiSilver Total", "=SUM('Employee Performance'!J2:J100)"),
        ("Overall Remarks", "All scheme goals achieved for the day.")
    ]
    for idx, (label, formula) in enumerate(scheme_summary, start=3):
        ws3.cell(row=idx, column=1, value=label).font = label_font
        val_cell = ws3.cell(row=idx, column=2, value=formula)
        val_cell.font = bold_font if "Total" in label else regular_font
        
        ws3.cell(row=idx, column=1).border = thin_border
        val_cell.border = thin_border
        
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 30
    
    # Save Workbook
    wb.save("templates/pothys_report_template.xlsx")
    print("Excel template created successfully at templates/pothys_report_template.xlsx")

create_pothys_template()
